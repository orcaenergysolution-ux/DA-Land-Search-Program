"""
NEM Generation Map builder.

Inputs (in same folder):
  - KCI Datafile Compiled NEM.xlsx
  - NEM Generation Information Oct 2025.xlsx
  - (optional, future) AEMO state map ingest

Output:
  - projects.json        unified project table
  - nem_generation_map.html   interactive Folium map

Stage logic (per brief):
  - In NEM Gen Info AND on AEMO map     -> use AEMO stage (Application/Committed)
  - In KCI AND NEM Gen Info, NOT on map -> Application
  - In KCI only                         -> Enquiry
  - NEM Gen Info Unit Status drives Existing / Committed / Anticipated buckets.
  - KCI NER type drives Enquiry vs Application when only in KCI.

We don't yet have the AEMO map extracted, so:
  - Treat NEM Gen Info "Committed" / "Anticipated" / "In Commissioning" as the
    AEMO-map "Committed" bucket.
  - Treat "Publicly Announced" as Application by default (later: if found on
    AEMO map, upgrade to Committed; if only in KCI, downgrade to Enquiry).
"""

from __future__ import annotations
import json
import re
import math
from pathlib import Path
import openpyxl

ROOT = Path(__file__).parent.parent
INPUTS = ROOT / "data" / "inputs"
INTERMEDIATE = ROOT / "data" / "intermediate"
OUTPUTS = ROOT / "outputs"
INTERMEDIATE.mkdir(parents=True, exist_ok=True)
OUTPUTS.mkdir(parents=True, exist_ok=True)
KCI_FILE = INPUTS / "KCI Datafile Compiled NEM202603201000.xlsx"
NEM_FILE = INPUTS / "NEM Generation Information Apr 2026.xlsx"

REGION_TO_STATE = {
    "NSW1": "NSW", "VIC1": "VIC", "QLD1": "QLD", "SA1": "SA", "TAS1": "TAS",
}

# ── Technology normalisation ──────────────────────────────────────────────────
# Maps raw Excel strings (lowercased, stripped) → clean display name.
# Prefix matches are tried in order; first match wins.
# Add new raw values here if new ones appear in the source files.
_TECH_PREFIX: list[tuple[str, str]] = [
    # Hybrid (check before solar/wind/battery so "battery; solar" → Hybrid)
    ("battery; solar",          "Hybrid"),
    ("battery; wind",           "Hybrid"),
    ("pv + bess",               "Hybrid"),
    ("solar + storage",         "Hybrid"),
    ("solar and bess",          "Hybrid"),
    # Solar
    ("solar thermal",           "Solar Thermal"),
    ("solar pv",                "Solar"),
    ("solar farm",              "Solar"),
    ("solarpv",                 "Solar"),
    ("solar",                   "Solar"),
    # Wind
    ("wind turbine - offshore", "Wind (Offshore)"),
    ("wind turbine - onshore",  "Wind (Onshore)"),
    ("wind turbine",            "Wind (Onshore)"),
    ("windturbine",             "Wind (Onshore)"),
    ("wind farm",               "Wind (Onshore)"),
    ("wind",                    "Wind (Onshore)"),
    # Storage
    ("storage - battery",       "Battery"),
    ("storage - pumped hydro",  "Pumped Hydro"),
    ("storage - pumped",        "Pumped Hydro"),
    ("storage",                 "Storage"),
    ("battery",                 "Battery"),
    # Hydro
    ("hydro - dam",             "Hydro"),
    ("hydro - run of river",    "Hydro"),
    ("hydro - pumped",          "Pumped Hydro"),
    ("hydro",                   "Hydro"),
    # Thermal / gas
    ("turbine - ocgt",          "OCGT"),
    ("turbine - ccgt",          "CCGT"),
    ("turbine - steam sub",     "Coal"),
    ("turbine - steam super",   "Coal"),
    ("turbine - other",         "Gas"),
    ("turbine",                 "Gas"),
    ("reciprocating engine",    "Gas"),
    ("gas",                     "Gas"),
    # Catch-all
    ("other - other",           "Other"),
    ("other",                   "Other"),
]


def normalise_tech(raw: str | None) -> str:
    """Map a raw technology string to a clean display category."""
    if not raw:
        return "Other"
    key = raw.strip().lower()
    for prefix, label in _TECH_PREFIX:
        if key.startswith(prefix):
            return label
    return raw.strip() or "Other"   # unrecognised — keep as-is so nothing is lost


_SOLAR   = {"Solar", "Solar Thermal"}
_WIND    = {"Wind (Onshore)", "Wind (Offshore)"}
_STORAGE = {"Battery", "Battery Storage", "Storage"}


def classify_hybrid_tech(techs: set[str]) -> str:
    """Return a combined technology label when a site has multiple technology types.

    Rules (checked in order):
      Solar + Storage only            → 'Solar and Storage'
      Wind + Storage only             → 'Wind and Storage'
      Anything else multi-tech        → 'Hybrid'
    """
    has_solar   = bool(techs & _SOLAR)
    has_wind    = bool(techs & _WIND)
    has_storage = bool(techs & _STORAGE)
    other = techs - _SOLAR - _WIND - _STORAGE  # Hydro, Gas, Coal, etc.

    if has_solar and has_storage and not has_wind and not other:
        return "Solar and Storage"
    if has_wind and has_storage and not has_solar and not other:
        return "Wind and Storage"
    return "Hybrid"


# AEMO Generation Information legend colours (approximate hex matches).
STAGE_COLOUR = {
    "Existing":      "#1f2937",  # dark grey / black
    "Retiring":      "#f97316",  # orange — still operating, announced closure
    "Commissioning": "#a21caf",  # purple — final testing before operation
    "Committed":     "#16a34a",  # green
    "Anticipated":   "#f59e0b",  # amber/yellow
    "Application":   "#2563eb",  # blue
    "Enquiry":       "#dc2626",  # red
    "Withdrawn":     "#9ca3af",  # light grey
    "Unknown":       "#6b7280",
}

STAGE_ORDER = ["Existing", "Retiring", "Commissioning", "Committed", "Anticipated", "Application", "Enquiry", "Withdrawn", "Unknown"]


def norm_name(s: str | None) -> str:
    if not s:
        return ""
    s = str(s).lower()
    s = re.sub(r"[​\xa0]", " ", s)             # zero-width / nbsp
    s = re.sub(r"\([^)]*\)", " ", s)                 # drop parenthetical
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return " ".join(s.split())


def to_float(x) -> float | None:
    if x is None or x == "":
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def read_nem_gen_info() -> list[dict]:
    wb = openpyxl.load_workbook(NEM_FILE, data_only=True, read_only=True)
    ws = wb["Generator Information"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 4:  # rows 1-3 are notes; row 4 (index 3) is header; data from row 5
            continue
        region = row[5]  # Region
        if not region or region not in REGION_TO_STATE:
            continue
        site = row[1]    # Site Name
        if not site:
            continue
        unit_status = (row[20] or "").strip().rstrip("*").strip()  # Commitment Status
        # collapse the weird "Committed*" / "Committed﻿" variants
        if unit_status.lower().startswith("committed"):
            unit_status = "Committed"
        # Agg Nameplate Capacity (MW AC), else Unit Capacity (MW AC), else MW DC
        cap = to_float(row[18]) or to_float(row[16]) or to_float(row[17])
        rows.append({
            "site_name": str(site).strip(),
            "region": region,
            "state": REGION_TO_STATE[region],
            "owner": (row[3] or "").strip() if row[3] else "",
            "technology": (row[9] or "").strip() if row[9] else "",   # Technology Type
            "fuel": (row[11] or "").strip() if row[11] else "",        # Gas Turbine Fuel Type
            "duid": (row[12] or "").strip() if row[12] else "",        # DUID
            "gen_unit_id": str(int(row[7])) if row[7] else "",         # Gen Info Unit ID (col 7)
            "kci_id": (row[2] or "").strip() if row[2] else "",        # AEMO KCI ID (col 2)
            "capacity_mw": cap,
            "storage_mwh": to_float(row[19]),   # Agg Nameplate Storage Capacity (MWh)
            "unit_status": unit_status,
            "asset_type": "Project",
            "_source": "NEM",
            "_key": norm_name(site),
        })
    return rows


def read_kci() -> list[dict]:
    wb = openpyxl.load_workbook(KCI_FILE, data_only=True, read_only=True)
    ws = wb["KCI Data"]

    # ── Pass 1: read all rows, skip only unidentifiable ones ─────────────────
    # col 1  = KCI datafile compilation date time stamp  (used in Pass 2)
    # col 6  = activity status  (Withdrawn / Cancelled / Active …)  (Pass 3)
    # col 10 = site name                                             (Pass 1 + 3)
    # col 12 = region  (NSW1 / VIC1 / QLD1 / SA1 / TAS1 …)         (Pass 3)
    # col 26 = AEMO KCI ID  (e.g. "V00077")                         (Pass 1 + 2)
    # Region, site name, and status filters are all deferred to Pass 3 so they
    # are evaluated against the LATEST revision of each project, not a stale one.
    # Only rows with no KCI ID AND no site name are dropped here — they cannot
    # be grouped or identified at all.
    raw: list[tuple] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:   # rows 0-1 = notes; row 2 = header; data from row 3
            continue
        kci_id = str(row[26]).strip() if row[26] else ""
        site   = str(row[10]).strip() if row[10] else ""
        if not kci_id and not site:   # ungroupable — nothing to identify this row
            continue
        raw.append(row)

    # ── Pass 2: deduplicate — keep latest row per AEMO KCI ID ────────────────
    # Group by AEMO KCI ID (col 26); fall back to (site_name, region) for rows
    # that have no KCI ID. Within each group, the row with the latest
    # compilation date (col 1) wins; ties are broken by raw row number so the
    # result is deterministic.
    from collections import defaultdict
    groups: dict[object, list[tuple[object, int, tuple]]] = defaultdict(list)
    for row_num, row in enumerate(raw):
        kci_id = str(row[26]).strip() if row[26] else None
        key    = kci_id if kci_id and kci_id.lower() not in ("none", "") \
                 else (str(row[10]).strip(), str(row[12]).strip())
        groups[key].append((row[1], row_num, row))   # (compile_date, row_num, row)

    deduped = [max(g, key=lambda x: (x[0] is not None, x[0], x[1]))[2]
               for g in groups.values()]

    dupes = len(raw) - len(deduped)
    if dupes:
        print(f"  [KCI] Deduplicated {dupes} older revision(s) "
              f"({len(raw)} rows -> {len(deduped)} unique KCI IDs)")

    # ── Pass 3: apply all filters on the latest revision of each project ──────
    # Evaluated here so we always use the most current state, not a stale one.
    before = len(deduped)
    deduped = [r for r in deduped
               if r[12] in REGION_TO_STATE                                  # NEM states only
               and (r[10] or "").strip()                                     # must have site name
               and (r[6] or "").strip().lower() not in {"withdrawn", "cancelled"}]
    dropped = before - len(deduped)
    if dropped:
        print(f"  [KCI] Dropped {dropped} row(s) after dedup "
              f"(non-NEM region / blank site / Withdrawn/Cancelled) "
              f"({before} -> {len(deduped)})")

    # ── Pass 3: build output dicts ────────────────────────────────────────────
    rows = []
    for row in deduped:
        ner = (row[3] or "").lower()
        if "application to connect" in ner:
            kci_stage = "Application"
        elif "connection enquiry" in ner:
            kci_stage = "Enquiry"
        else:
            kci_stage = "Enquiry"
        cap_upper = to_float(row[14]) or to_float(row[25]) or to_float(row[13])
        rows.append({
            "site_name":       str(row[10]).strip().replace("​", "").replace("\xa0", " ").strip(),
            "region":          row[12],
            "state":           REGION_TO_STATE[row[12]],
            "owner":           (row[7] or "").strip() if row[7] else "",
            "technology":      (row[17] or "").strip() if row[17] else "",
            "fuel":            "",
            "duid":            "",
            "capacity_mw":     cap_upper,
            "storage_mwh":     None,
            "location_desc":   (row[11] or "").strip() if row[11] else "",
            "ner_type":        row[3] or "",
            "activity_status": (row[6] or "").strip(),
            "kci_stage":       kci_stage,
            "kci_id":          str(row[26]).strip() if row[26] else "",
            "compile_date":    str(row[1]) if row[1] else "",
            "_source":         "KCI",
            "_key":            norm_name(str(row[10])),
        })
    return rows


def load_aemo_map_names() -> set[str]:
    """Return normalized site-name keys that appear on any AEMO map PDF."""
    p = INTERMEDIATE / "aemo_pdf_matches.json"
    if not p.exists():
        return set()
    data = json.loads(p.read_text(encoding="utf-8"))
    keys: set[str] = set()
    for _, info in data.items():
        for m in info.get("matches", []):
            keys.add(norm_name(m["matched_site"]))
    return keys


def merge(nem: list[dict], kci: list[dict]) -> list[dict]:
    """Merge NEM and KCI records.

    Matching priority:
      1. AEMO KCI ID  — NEM col 2 matches KCI col 26 (exact, reliable)
      2. Normalised site name — fallback for NEM rows with no KCI ID

    Grouping by (key, unit_status) prevents two classes of bad summation:
      1. Same tech, different status  -> separate projects (e.g. Eraring Existing vs Committed)
      2. Different tech, same status  -> hybrid label   (e.g. Solar+Battery -> 'Solar and Storage')
    Within each (key, status) group, capacities of the same unit type are summed as usual.
    """
    # Group NEM rows by (norm-key, unit_status) — one merged entry per combination
    nem_by_key_status: dict[tuple, list[dict]] = {}
    for r in nem:
        k = (r["_key"], r["unit_status"])
        nem_by_key_status.setdefault(k, []).append(r)

    # Build KCI lookups — by ID (preferred) and by normalised name (fallback)
    kci_by_id:  dict[str, dict]        = {}   # kci_id  -> single KCI record (already deduped)
    kci_by_key: dict[str, list[dict]]  = {}   # _key    -> list of KCI records
    for r in kci:
        kid = (r.get("kci_id") or "").strip()
        if kid:
            kci_by_id[kid] = r
        kci_by_key.setdefault(r["_key"], []).append(r)

    aemo_keys = load_aemo_map_names()
    merged: list[dict] = []
    seen_kci_ids:  set[str] = set()   # KCI IDs already matched to a NEM record
    seen_nem_keys: set[str] = set()   # NEM name keys already processed

    matched_by_id   = 0
    matched_by_name = 0

    # ── NEM pass (NEM is authoritative for capacity / DUID) ──────────────────
    for (key, _status), group in nem_by_key_status.items():
        if not key:
            continue
        seen_nem_keys.add(key)
        primary   = max(group, key=lambda r: r.get("capacity_mw") or 0)
        total_cap = sum((r.get("capacity_mw") or 0) for r in group)
        storage_vals = [r["storage_mwh"] for r in group if r.get("storage_mwh")]
        storage_mwh  = round(sum(storage_vals), 2) if storage_vals else None

        # Classify technology — hybrid label when multiple distinct techs share the same status
        techs      = {normalise_tech(r["technology"]) for r in group}
        technology = classify_hybrid_tech(techs) if len(techs) > 1 else next(iter(techs))

        # Match KCI: ID first, name fallback
        nem_kci_id = (primary.get("kci_id") or "").strip()
        if nem_kci_id and nem_kci_id in kci_by_id:
            kci_match = [kci_by_id[nem_kci_id]]
            seen_kci_ids.add(nem_kci_id)
            matched_by_id += 1
        else:
            kci_match = kci_by_key.get(key, [])
            if kci_match:
                matched_by_name += 1

        kci_id_out   = kci_match[0].get("kci_id", "") if kci_match else nem_kci_id
        loc          = next((k["location_desc"] for k in kci_match if k.get("location_desc")), "")
        on_aemo      = key in aemo_keys
        gen_unit_ids = sorted({r["gen_unit_id"] for r in group if r.get("gen_unit_id")})

        stage = classify_stage(primary, in_kci=bool(kci_match), on_aemo_map=on_aemo)
        merged.append({
            "site_name":    primary["site_name"],
            "region":       primary["region"],
            "state":        primary["state"],
            "owner":        primary["owner"],
            "technology":   technology,
            "fuel":         primary["fuel"],
            "capacity_mw":  round(total_cap, 2) if total_cap else primary["capacity_mw"],
            "storage_mwh":  storage_mwh,
            "unit_status":  primary["unit_status"],
            "asset_type":   primary["asset_type"],
            "location_desc": loc,
            "stage":        stage,
            "source":       "NEM+KCI" if kci_match else "NEM",
            "on_aemo_map":  on_aemo,
            "duid":         primary.get("duid", ""),
            "kci_id":       kci_id_out,
            "gen_unit_ids": gen_unit_ids,
        })

    print(f"  [merge] NEM+KCI matches: {matched_by_id} by ID, {matched_by_name} by name")

    # ── KCI-only pass (Enquiry / Application if on AEMO map) ─────────────────
    for key, group in kci_by_key.items():
        if not key:
            continue
        if key in seen_nem_keys:
            continue
        primary = max(group, key=lambda r: r.get("capacity_mw") or 0)
        kid = (primary.get("kci_id") or "").strip()
        if kid and kid in seen_kci_ids:
            continue   # already linked to a NEM record by ID match
        on_aemo = key in aemo_keys
        # KCI-only projects are always Enquiry unless they appear on the AEMO map,
        # in which case they are Application. NEM Gen Info drives Application and above.
        stage = "Application" if on_aemo else "Enquiry"
        merged.append({
            "site_name":    primary["site_name"],
            "region":       primary["region"],
            "state":        primary["state"],
            "owner":        primary["owner"],
            "technology":   normalise_tech(primary["technology"]),
            "fuel":         "",
            "capacity_mw":  primary["capacity_mw"],
            "storage_mwh":  None,
            "unit_status":  "",
            "asset_type":   "Project",
            "location_desc": primary.get("location_desc", ""),
            "stage":        stage,
            "source":       "KCI",
            "on_aemo_map":  on_aemo,
            "duid":         "",
            "kci_id":       kid,
            "gen_unit_ids": [],
        })

    return merged


def classify_stage(nem_row: dict, in_kci: bool, on_aemo_map: bool = False) -> str:
    """Map NEM Gen Info Unit Status (+ AEMO map presence) to a display stage.

    Rules:
      In Service                        -> Existing
      Announced Withdrawal              -> Retiring  (still operating, closure date notified)
      Committed / In Commissioning      -> Committed
      Anticipated                       -> Anticipated
      Withdrawn - Permanent             -> Withdrawn
      Publicly Announced                -> Application
    """
    us = (nem_row.get("unit_status") or "").lower()
    if us.startswith("in service"):
        return "Existing"
    if "announced withdrawal" in us:
        return "Retiring"
    if "commissioning" in us:
        return "Commissioning"
    if us.startswith("committed"):
        return "Committed"
    if "anticipated" in us:
        return "Anticipated"
    if "withdrawn" in us:
        return "Withdrawn"
    if "publicly announced" in us or us == "proposed":
        return "Application"
    return "Unknown"




def main() -> None:
    nem = read_nem_gen_info()
    kci = read_kci()
    merged = merge(nem, kci)
    # Initialise lat/lon to None — geocode.py fills these in
    for r in merged:
        r.setdefault("lat", None)
        r.setdefault("lon", None)

    # Stats
    from collections import Counter
    stage_ct  = Counter(r["stage"]      for r in merged)
    source_ct = Counter(r["source"]     for r in merged)
    region_ct = Counter(r["region"]     for r in merged)
    tech_ct   = Counter(r["technology"] for r in merged)
    hybrid_techs = {"Solar and Storage", "Wind and Storage", "Hybrid"}
    hybrid_ct = {t: tech_ct[t] for t in sorted(hybrid_techs) if tech_ct.get(t)}
    print(f"Total projects: {len(merged)}")
    print(f"By stage: {dict(stage_ct)}")
    print(f"By source: {dict(source_ct)}")
    print(f"By region: {dict(region_ct)}")
    if hybrid_ct:
        print(f"Hybrid/combined techs: {hybrid_ct}")

    (INTERMEDIATE / "projects.json").write_text(
        json.dumps(merged, indent=2, default=str), encoding="utf-8"
    )
    print(f"Wrote {INTERMEDIATE/'projects.json'}")



if __name__ == "__main__":
    main()
